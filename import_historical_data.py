#!/usr/bin/env python3
"""
历史数据导入中心 — 2021-2025赛季欧洲五大联赛+欧冠+欧联 Excel 数据导入流水线

功能:
  1. 自动读取所有 sheet，智能识别两种列布局
  2. 解析多层表头，提取博彩公司名称
  3. 清洗为标准数据库结构
  4. 导入 SQLite + 导出 CSV/JSON
  5. 自动去重
  6. 生成数据质量报告
  7. 为 AI/回测准备特征字段

用法:
  python import_historical_data.py
  python import_historical_data.py --excel "path/to/file.xlsx"
  python import_historical_data.py --skip-json --skip-csv
"""

import hashlib
import json
import os
import re
import sqlite3
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from config.paths import RAW_DATA_DIR, DATABASE_DIR, DB_FOOTBALL_HISTORY

# ── 项目根目录 ──────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent
DATA_RAW = RAW_DATA_DIR
DATA_PROCESSED = RAW_DATA_DIR  # processed data stays in raw_data
DATA_DB = DB_FOOTBALL_HISTORY
DATA_REPORTS = PROJECT_ROOT / "data" / "reports"

# ── Excel 文件路径 ──────────────────────────────────────────
EXCEL_FILE = DATA_RAW / "五大联赛+欧冠+欧联足球比赛数据统计2021-2025赛季(1).xlsx"

# ── Sheet → 联赛映射 ────────────────────────────────────────
SHEET_LEAGUE_MAP = {
    "英超": {"name": "English Premier League", "country": "England", "code": "EPL"},
    "德甲": {"name": "German Bundesliga", "country": "Germany", "code": "BUN"},
    "意甲": {"name": "Italian Serie A", "country": "Italy", "code": "ISA"},
    "西甲": {"name": "Spanish La Liga", "country": "Spain", "code": "LLG"},
    "法甲": {"name": "French Ligue 1", "country": "France", "code": "FL1"},
    "欧冠": {"name": "UEFA Champions League", "country": "Europe", "code": "UCL"},
    "欧罗巴": {"name": "UEFA Europa League", "country": "Europe", "code": "UEL"},
    # "汇总" sheet — 包含所有联赛数据，应最后处理（去重后自动跳过）
    "汇总": {"name": "Summary", "country": "Unknown", "code": "SUM"},
}

# ── 博彩公司名称映射（根据 Excel Row1 中的缩写） ──────────────
BOOKMAKER_GROUPS = {
    "36*": "Bet365",
    "澳*": "Macau",
    "必*": "Betfair",
    "皇*": "Crown",
    "立*": "Ladbrokes",
    "威*": "William Hill",
}

# ── 欧洲赔率博彩公司（6家，每家有初盘+即时盘） ─────────────────
EUROPE_BOOKMAKERS = ["36*", "澳*", "必*", "皇*", "立*", "威*"]

# ── 亚盘/大小球博彩公司（5家，每家有初盘+终盘） ──────────────
ASIAN_BOOKMAKERS = ["36*", "澳*", "皇*", "立*", "威*"]


def safe_float(val):
    """安全的浮点数转换"""
    if val is None or val == "" or val == "None":
        return None
    try:
        return float(str(val).strip().replace(",", "."))
    except (ValueError, TypeError):
        return None


def safe_int(val):
    """安全的整数转换"""
    if val is None or val == "" or val == "None":
        return None
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return None


def safe_str(val):
    """安全的字符串转换并清理"""
    if val is None:
        return ""
    s = str(val).strip()
    # 清理方括号排名标记如 [10]
    s = re.sub(r"^\[\d+\]\s*", "", s)
    return s


def normalize_team_name(name: str) -> str:
    """标准化球队名称"""
    name = safe_str(name)
    # 去掉排名标记
    name = re.sub(r"^\[\d+\]", "", name).strip()
    return name


def parse_score(score_str: str):
    """解析比分字符串，返回 (home, away)"""
    if not score_str:
        return None, None
    s = str(score_str).strip().replace(" ", "")
    # 格式: "2-3" 或 "1-0"
    match = re.match(r"(\d+)\s*-\s*(\d+)", s)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None, None


def parse_kickoff_time(season_str: str, time_str: str):
    """解析开球时间，返回 ISO 格式字符串"""
    if not time_str:
        return None
    parts = str(time_str).strip().split()
    if len(parts) >= 1:
        date_part = parts[0]  # "08-14"
        time_part = parts[1] if len(parts) >= 2 else "00:00"
        try:
            season_year = int(str(season_str).strip()[:4])
        except ValueError:
            season_year = 2021
        # 判断月份：8月以后的属于赛季年份，1-7月属于下一年
        month_str = date_part.split("-")[0]
        try:
            month = int(month_str)
        except ValueError:
            month = 1
        if month >= 8:
            year = season_year
        else:
            year = season_year + 1
        try:
            return f"{year}-{date_part} {time_part}:00"
        except Exception:
            return None
    return None


def generate_id(*parts: str) -> str:
    """生成唯一 ID"""
    raw = "|".join(p.strip().lower() for p in parts if p)
    return hashlib.md5(raw.encode()).hexdigest()[:12]


def detect_layout_type(headers_row2: list, max_col: int) -> str:
    """
    检测列布局类型。
    Type A (法甲/德甲/意甲/西甲/欧冠/欧洲杯):
      cols 1-10: basic info, cols 11-46: euro odds, cols 47-76: asian,
      cols 77-106: over/under, cols 107-111: results
    Type B (英超):
      cols 1-10: basic info, cols 11-15: results, cols 16-51: euro odds,
      cols 52-81: asian, cols 82-111: over/under
    """
    # 检查 col 11 的 header: Type B 的第11列是"半场比分"(result),
    # Type A 的第11列是"胜"(odds)
    if max_col >= 11:
        h11 = headers_row2[10] if len(headers_row2) > 10 else ""
        # "半场" contains 半场 chars
        if "半" in h11 or "比" in h11:
            return "B"
    return "A"


class HistoricalDataImporter:
    """历史数据导入器"""

    def __init__(self, excel_path: str = None):
        self.excel_path = Path(excel_path or EXCEL_FILE)
        self.db_path = DATA_DB
        self.processed_dir = DATA_PROCESSED
        self.reports_dir = DATA_REPORTS

        # 确保目录存在
        self.processed_dir.mkdir(parents=True, exist_ok=True)
        self.reports_dir.mkdir(parents=True, exist_ok=True)

        # 统计信息
        self.stats = {
            "total_rows": 0,
            "valid_matches": 0,
            "skipped_empty": 0,
            "duplicates": 0,
            "euro_odds_count": 0,
            "asian_odds_count": 0,
            "over_under_odds_count": 0,
            "errors": [],
            "per_sheet": {},
            "teams_count": 0,
            "leagues_count": 0,
            "seasons": set(),
        }

        # 内存缓存
        self.leagues_cache: dict[str, dict] = {}
        self.teams_cache: dict[str, dict] = {}
        self._seen_match_ids: set = set()

        # 数据库连接
        self.conn: Optional[sqlite3.Connection] = None

    # ── 数据库设置 ──────────────────────────────────────────

    def setup_database(self):
        """创建 SQLite 数据库及表结构"""
        self.conn = sqlite3.connect(str(self.db_path))
        self.conn.execute("PRAGMA journal_mode=WAL")
        self.conn.execute("PRAGMA foreign_keys=ON")

        self.conn.executescript("""
        CREATE TABLE IF NOT EXISTS leagues (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            name_cn TEXT,
            country TEXT,
            code TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS teams (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            name_cn TEXT,
            league_id TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (league_id) REFERENCES leagues(id)
        );

        CREATE TABLE IF NOT EXISTS seasons (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            start_year INTEGER,
            end_year INTEGER
        );

        CREATE TABLE IF NOT EXISTS matches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            match_id TEXT UNIQUE NOT NULL,
            source TEXT DEFAULT 'historical_excel',
            league_id TEXT,
            season TEXT,
            home_team_id TEXT,
            home_team TEXT NOT NULL,
            away_team_id TEXT,
            away_team TEXT NOT NULL,
            kickoff_time TEXT,
            home_score INTEGER,
            away_score INTEGER,
            score_display TEXT,
            half_time_score TEXT,
            total_goals INTEGER,
            half_full_result TEXT,
            ft_result TEXT,
            round TEXT,
            status TEXT DEFAULT 'finished',
            home_ranking INTEGER,
            away_ranking INTEGER,
            collected_at TEXT DEFAULT (datetime('now')),
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (league_id) REFERENCES leagues(id),
            FOREIGN KEY (home_team_id) REFERENCES teams(id),
            FOREIGN KEY (away_team_id) REFERENCES teams(id)
        );

        CREATE TABLE IF NOT EXISTS odds_europe (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            match_id TEXT NOT NULL,
            bookmaker TEXT NOT NULL,
            odds_type TEXT NOT NULL CHECK(odds_type IN ('opening', 'closing')),
            odds_home REAL,
            odds_draw REAL,
            odds_away REAL,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (match_id) REFERENCES matches(match_id),
            UNIQUE(match_id, bookmaker, odds_type)
        );

        CREATE TABLE IF NOT EXISTS odds_asian (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            match_id TEXT NOT NULL,
            bookmaker TEXT NOT NULL,
            odds_type TEXT NOT NULL CHECK(odds_type IN ('opening', 'closing')),
            high_water REAL,
            handicap TEXT,
            low_water REAL,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (match_id) REFERENCES matches(match_id),
            UNIQUE(match_id, bookmaker, odds_type)
        );

        CREATE TABLE IF NOT EXISTS odds_over_under (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            match_id TEXT NOT NULL,
            bookmaker TEXT NOT NULL,
            odds_type TEXT NOT NULL CHECK(odds_type IN ('opening', 'closing')),
            over_water REAL,
            handicap TEXT,
            under_water REAL,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (match_id) REFERENCES matches(match_id),
            UNIQUE(match_id, bookmaker, odds_type)
        );

        CREATE TABLE IF NOT EXISTS odds_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            match_id TEXT NOT NULL,
            bookmaker TEXT NOT NULL,
            odds_home REAL,
            odds_draw REAL,
            odds_away REAL,
            asian_handicap TEXT,
            over_under_handicap TEXT,
            snapshot_at TEXT NOT NULL DEFAULT (datetime('now')),
            FOREIGN KEY (match_id) REFERENCES matches(match_id)
        );

        CREATE TABLE IF NOT EXISTS import_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_time TEXT DEFAULT (datetime('now')),
            source_file TEXT,
            source_sheet TEXT,
            total_rows INTEGER,
            valid_rows INTEGER,
            skipped_rows INTEGER,
            duplicate_rows INTEGER,
            errors TEXT,
            duration_seconds REAL
        );

        CREATE TABLE IF NOT EXISTS match_features (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            match_id TEXT UNIQUE NOT NULL,
            -- 欧赔隐含概率 (Bet365 closing)
            implied_home_prob REAL,
            implied_draw_prob REAL,
            implied_away_prob REAL,
            -- 欧赔变化 (opening -> closing)
            odds_movement_home REAL,
            odds_movement_draw REAL,
            odds_movement_away REAL,
            -- 亚盘变化
            asian_handicap_open TEXT,
            asian_handicap_close TEXT,
            asian_movement_flag INTEGER DEFAULT 0,
            -- 大小球
            over_under_line REAL,
            over_prob_implied REAL,
            -- 赔率离散度 (多家博彩公司标准差)
            odds_std_home REAL,
            odds_std_draw REAL,
            odds_std_away REAL,
            -- 标签（用于 AI）
            home_win INTEGER DEFAULT 0,
            draw INTEGER DEFAULT 0,
            away_win INTEGER DEFAULT 0,
            total_goals INTEGER,
            over_2_5 INTEGER DEFAULT 0,
            both_teams_score INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (match_id) REFERENCES matches(match_id)
        );

        -- 索引
        CREATE INDEX IF NOT EXISTS idx_matches_league ON matches(league_id);
        CREATE INDEX IF NOT EXISTS idx_matches_season ON matches(season);
        CREATE INDEX IF NOT EXISTS idx_matches_kickoff ON matches(kickoff_time);
        CREATE INDEX IF NOT EXISTS idx_matches_home_team ON matches(home_team);
        CREATE INDEX IF NOT EXISTS idx_matches_away_team ON matches(away_team);
        CREATE INDEX IF NOT EXISTS idx_matches_ft_result ON matches(ft_result);
        CREATE INDEX IF NOT EXISTS idx_odds_europe_match ON odds_europe(match_id);
        CREATE INDEX IF NOT EXISTS idx_odds_asian_match ON odds_asian(match_id);
        CREATE INDEX IF NOT EXISTS idx_odds_over_under_match ON odds_over_under(match_id);
        CREATE INDEX IF NOT EXISTS idx_teams_name ON teams(name);
        """)

    # ── 联赛和球队管理 ──────────────────────────────────────

    def get_or_create_league(self, sheet_name: str) -> str:
        """获取或创建联赛记录"""
        league_info = SHEET_LEAGUE_MAP.get(sheet_name, {
            "name": sheet_name, "country": "Unknown", "code": sheet_name[:3].upper()
        })
        league_id = generate_id(league_info["name"])

        if league_id not in self.leagues_cache:
            self.leagues_cache[league_id] = {
                "id": league_id,
                "name": league_info["name"],
                "name_cn": sheet_name,
                "country": league_info["country"],
                "code": league_info["code"],
            }
        return league_id

    def get_or_create_team(self, team_name: str, league_id: str = None) -> str:
        """获取或创建球队记录"""
        team_name_clean = normalize_team_name(team_name)
        if not team_name_clean:
            return None
        team_id = generate_id(team_name_clean)

        if team_id not in self.teams_cache:
            self.teams_cache[team_id] = {
                "id": team_id,
                "name": team_name_clean,
                "name_cn": team_name_clean,
                "league_id": league_id,
            }
        return team_id

    # ── Excel 解析 ─────────────────────────────────────────

    def parse_excel(self):
        """解析 Excel 文件的所有 sheet"""
        import openpyxl

        print(f"\n{'='*60}")
        print(f"开始解析 Excel 文件: {self.excel_path.name}")
        print(f"文件大小: {os.path.getsize(self.excel_path) / 1024 / 1024:.2f} MB")
        print(f"{'='*60}")

        wb = openpyxl.load_workbook(str(self.excel_path), read_only=True, data_only=True)
        all_matches = []
        all_euro_odds = []
        all_asian_odds = []
        all_over_under_odds = []

        # 将 "汇总" sheet 放到最后处理，避免覆盖各联赛的 league_id
        sheet_names = list(wb.sheetnames)
        summary_sheets = [s for s in sheet_names if "汇总" in str(s) or "总" in str(s)]
        normal_sheets = [s for s in sheet_names if s not in summary_sheets]
        ordered_sheets = normal_sheets + summary_sheets
        if summary_sheets:
            print(f"检测到汇总 sheet: {summary_sheets}, 将在最后处理")

        for sheet_name in ordered_sheets:
            print(f"\n--- 处理 Sheet: {sheet_name} ---")
            ws = wb[sheet_name]
            sheet_start = time.time()

            # 读取前两行（表头）
            row1 = []
            row2 = []
            for cell in ws[1]:
                row1.append(str(cell.value) if cell.value is not None else "")
            for cell in ws[2]:
                row2.append(str(cell.value) if cell.value is not None else "")

            # 补齐列数
            while len(row1) < ws.max_column:
                row1.append("")
            while len(row2) < ws.max_column:
                row2.append("")

            # 检测布局类型
            layout = detect_layout_type(row2, ws.max_column)
            print(f"  布局类型: Type {layout} ({'英超-结果在赔率前' if layout == 'B' else '标准-结果在赔率后'})")

            # 解析博彩公司名称
            bookmaker_names = self._parse_bookmaker_names(row1, row2)
            print(f"  识别博彩公司: {list(bookmaker_names.keys())}")

            # 计数
            sheet_valid = 0
            sheet_skipped = 0
            league_id = self.get_or_create_league(sheet_name)

            # 解析数据行（从第3行开始）
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
                self.stats["total_rows"] += 1
                row_list = list(row)
                # 补齐列
                while len(row_list) < ws.max_column:
                    row_list.append(None)

                # 检查是否空行
                if row_list[0] is None or str(row_list[0]).strip() == "":
                    sheet_skipped += 1
                    self.stats["skipped_empty"] += 1
                    continue

                try:
                    match_data, euro_odds, asian_odds, ou_odds = self._parse_row(
                        row_list, layout, league_id, sheet_name, bookmaker_names
                    )
                    if match_data:
                        all_matches.append(match_data)
                        all_euro_odds.extend(euro_odds)
                        all_asian_odds.extend(asian_odds)
                        all_over_under_odds.extend(ou_odds)
                        sheet_valid += 1
                        self.stats["valid_matches"] += 1
                except Exception as e:
                    self.stats["errors"].append(f"Sheet={sheet_name} Row={row_idx}: {e}")

            sheet_elapsed = time.time() - sheet_start
            self.stats["per_sheet"][sheet_name] = {
                "rows": ws.max_row - 2,
                "valid": sheet_valid,
                "skipped": sheet_skipped,
                "seconds": sheet_elapsed,
            }
            print(f"  有效: {sheet_valid}, 跳过: {sheet_skipped}, 耗时: {sheet_elapsed:.1f}s")

        wb.close()
        return all_matches, all_euro_odds, all_asian_odds, all_over_under_odds

    def _parse_bookmaker_names(self, row1: list, row2: list) -> dict:
        """解析博彩公司名称映射（从双层表头中提取）"""
        # 从 Row1 中提取唯一的博彩公司分组名
        names = {}
        for val in row1:
            if val and val not in names and val != "None":
                # 清理名称
                clean = val.strip()
                if clean:
                    base_name = BOOKMAKER_GROUPS.get(clean, clean)
                    names[clean] = base_name
        return names

    def _parse_row(self, row: list, layout: str, league_id: str,
                   sheet_name: str, bookmaker_names: dict):
        """解析单行数据"""
        # ── 基础信息 (cols 1-10, 两种布局通用) ──────────────
        # row[0]=序号, row[1]=赛季, row[2]=联赛, row[3]=轮次,
        # row[4]=比赛时间, row[5]=状态, row[6]=排名, row[7]=主队,
        # row[8]=客队, row[9]=排名
        season = safe_str(row[1])
        league_name_in_sheet = safe_str(row[2])
        round_name = safe_str(row[3])
        kickoff_raw = safe_str(row[4])
        status_raw = safe_str(row[5])
        home_ranking = safe_int(row[6]) if row[6] else None
        home_team = safe_str(row[7])
        away_team = safe_str(row[8])
        away_ranking = safe_int(row[9]) if row[9] else None

        if not home_team or not away_team:
            return None, [], [], []

        # ── 开球时间 ─────────────────────────────────────────
        kickoff_time = parse_kickoff_time(season, kickoff_raw)

        # ── 根据布局读取比赛结果 ─────────────────────────────
        if layout == "B":
            # 英超：结果在 cols 11-15 (0-indexed: 10-14)
            half_time = safe_str(row[10])
            full_time = safe_str(row[11])
            total_goals = safe_int(row[12])
            half_full = safe_str(row[13])
            ft_result = safe_str(row[14])
            euro_offset = 15  # 欧赔从 col 16 (0-indexed: 15) 开始
        else:
            # 标准布局：欧赔从 col 11 (0-indexed: 10) 开始
            euro_offset = 10
            # 结果在固定位置：O/U 结束后的 107-111 列 (0-indexed: 106-110)
            # 欧赔 36 列 + 亚盘 30 列 + 大小球 30 列 = 96 列，从 col 11 开始
            # 结果起始 = 10 + 96 = 106
            result_start = euro_offset + 36 + 30 + 30  # = 106
            half_time = safe_str(row[result_start]) if len(row) > result_start else ""
            full_time = safe_str(row[result_start + 1]) if len(row) > result_start + 1 else ""
            total_goals = safe_int(row[result_start + 2]) if len(row) > result_start + 2 else None
            half_full = safe_str(row[result_start + 3]) if len(row) > result_start + 3 else ""
            ft_result = safe_str(row[result_start + 4]) if len(row) > result_start + 4 else ""

        # ── 比分解析 ─────────────────────────────────────────
        home_score, away_score = parse_score(full_time)
        score_display = full_time.strip() if full_time else ""

        # ── 生成 match_id ────────────────────────────────────
        match_id = generate_id("historical_excel", home_team, away_team,
                               kickoff_time or "", season)

        # ── 球队 ID ──────────────────────────────────────────
        home_team_id = self.get_or_create_team(home_team, league_id)
        away_team_id = self.get_or_create_team(away_team, league_id)

        # ── 赛季名称和年份 ────────────────────────────────────
        try:
            start_year = int(season[:4])
        except ValueError:
            start_year = None

        self.stats["seasons"].add(season)

        # ── 构建比赛记录 ─────────────────────────────────────
        match_data = {
            "match_id": match_id,
            "source": "historical_excel",
            "league_id": league_id,
            "season": season,
            "home_team_id": home_team_id,
            "home_team": normalize_team_name(home_team),
            "away_team_id": away_team_id,
            "away_team": normalize_team_name(away_team),
            "kickoff_time": kickoff_time,
            "home_score": home_score,
            "away_score": away_score,
            "score_display": score_display,
            "half_time_score": half_time.strip() if half_time else "",
            "total_goals": total_goals,
            "half_full_result": half_full.strip() if half_full else "",
            "ft_result": ft_result.strip() if ft_result else "",
            "round": round_name,
            "status": "finished" if status_raw in ("完", "Finished", "") else status_raw,
            "home_ranking": home_ranking,
            "away_ranking": away_ranking,
            "collected_at": datetime.now().isoformat(),
        }

        # ── 欧洲赔率 ─────────────────────────────────────────
        euro_odds = self._parse_euro_odds(row, euro_offset, match_id)

        # ── 亚盘 ─────────────────────────────────────────────
        asian_offset = euro_offset + 36  # 欧洲赔率占 36 列
        asian_odds = self._parse_asian_odds(row, asian_offset, match_id)

        # ── 大小球 ───────────────────────────────────────────
        ou_offset = asian_offset + 30  # 亚盘占 30 列
        over_under_odds = self._parse_over_under_odds(row, ou_offset, match_id)

        return match_data, euro_odds, asian_odds, over_under_odds

    def _parse_euro_odds(self, row: list, offset: int, match_id: str) -> list:
        """解析欧洲赔率（12组 × 3列 = 36列）"""
        results = []
        bookmaker_order = EUROPE_BOOKMAKERS
        for i, bk in enumerate(bookmaker_order):
            # 初盘 (opening)
            base_open = offset + i * 6
            h_open = safe_float(row[base_open]) if base_open < len(row) else None
            d_open = safe_float(row[base_open + 1]) if base_open + 1 < len(row) else None
            a_open = safe_float(row[base_open + 2]) if base_open + 2 < len(row) else None

            if h_open or d_open or a_open:
                results.append({
                    "match_id": match_id,
                    "bookmaker": BOOKMAKER_GROUPS.get(bk, bk),
                    "bookmaker_raw": bk,
                    "odds_type": "opening",
                    "odds_home": h_open,
                    "odds_draw": d_open,
                    "odds_away": a_open,
                })

            # 即时盘 (closing)
            base_close = offset + i * 6 + 3
            h_close = safe_float(row[base_close]) if base_close < len(row) else None
            d_close = safe_float(row[base_close + 1]) if base_close + 1 < len(row) else None
            a_close = safe_float(row[base_close + 2]) if base_close + 2 < len(row) else None

            if h_close or d_close or a_close:
                results.append({
                    "match_id": match_id,
                    "bookmaker": BOOKMAKER_GROUPS.get(bk, bk),
                    "bookmaker_raw": bk,
                    "odds_type": "closing",
                    "odds_home": h_close,
                    "odds_draw": d_close,
                    "odds_away": a_close,
                })

        return results

    def _parse_asian_odds(self, row: list, offset: int, match_id: str) -> list:
        """解析亚盘赔率（5家 × 2类型 × 3列 = 30列）"""
        results = []
        for i, bk in enumerate(ASIAN_BOOKMAKERS):
            # 初盘 (opening)
            base_open = offset + i * 6
            high_w = safe_float(row[base_open]) if base_open < len(row) else None
            handicap = safe_str(row[base_open + 1]) if base_open + 1 < len(row) else ""
            low_w = safe_float(row[base_open + 2]) if base_open + 2 < len(row) else None

            if high_w or handicap or low_w:
                results.append({
                    "match_id": match_id,
                    "bookmaker": BOOKMAKER_GROUPS.get(bk, bk),
                    "bookmaker_raw": bk,
                    "odds_type": "opening",
                    "high_water": high_w,
                    "handicap": handicap,
                    "low_water": low_w,
                })

            # 终盘 (closing)
            base_close = offset + i * 6 + 3
            high_w_c = safe_float(row[base_close]) if base_close < len(row) else None
            handicap_c = safe_str(row[base_close + 1]) if base_close + 1 < len(row) else ""
            low_w_c = safe_float(row[base_close + 2]) if base_close + 2 < len(row) else None

            if high_w_c or handicap_c or low_w_c:
                results.append({
                    "match_id": match_id,
                    "bookmaker": BOOKMAKER_GROUPS.get(bk, bk),
                    "bookmaker_raw": bk,
                    "odds_type": "closing",
                    "high_water": high_w_c,
                    "handicap": handicap_c,
                    "low_water": low_w_c,
                })

        return results

    def _parse_over_under_odds(self, row: list, offset: int, match_id: str) -> list:
        """解析大小球赔率（5家 × 2类型 × 3列 = 30列）"""
        results = []
        for i, bk in enumerate(ASIAN_BOOKMAKERS):
            # 初盘 (opening)
            base_open = offset + i * 6
            over_w = safe_float(row[base_open]) if base_open < len(row) else None
            handicap = safe_str(row[base_open + 1]) if base_open + 1 < len(row) else ""
            under_w = safe_float(row[base_open + 2]) if base_open + 2 < len(row) else None

            if over_w or handicap or under_w:
                results.append({
                    "match_id": match_id,
                    "bookmaker": BOOKMAKER_GROUPS.get(bk, bk),
                    "bookmaker_raw": bk,
                    "odds_type": "opening",
                    "over_water": over_w,
                    "handicap": handicap,
                    "under_water": under_w,
                })

            # 终盘 (closing)
            base_close = offset + i * 6 + 3
            over_w_c = safe_float(row[base_close]) if base_close < len(row) else None
            handicap_c = safe_str(row[base_close + 1]) if base_close + 1 < len(row) else ""
            under_w_c = safe_float(row[base_close + 2]) if base_close + 2 < len(row) else None

            if over_w_c or handicap_c or under_w_c:
                results.append({
                    "match_id": match_id,
                    "bookmaker": BOOKMAKER_GROUPS.get(bk, bk),
                    "bookmaker_raw": bk,
                    "odds_type": "closing",
                    "over_water": over_w_c,
                    "handicap": handicap_c,
                    "under_water": under_w_c,
                })

        return results

    # ── 数据库写入 ──────────────────────────────────────────

    def save_to_database(self, matches, euro_odds, asian_odds, over_under_odds):
        """写入 SQLite 数据库，自动去重"""
        print(f"\n{'='*60}")
        print("写入数据库...")
        print(f"{'='*60}")

        # ── 联赛 ─────────────────────────────────────────────
        league_sql = """
            INSERT OR IGNORE INTO leagues (id, name, name_cn, country, code)
            VALUES (?, ?, ?, ?, ?)
        """
        league_count = 0
        for lid, info in self.leagues_cache.items():
            c = self.conn.execute(league_sql, (
                info["id"], info["name"], info.get("name_cn", ""),
                info.get("country", ""), info.get("code", ""),
            ))
            if c.rowcount > 0:
                league_count += 1
        self.stats["leagues_count"] = league_count
        print(f"  联赛: {league_count} 条 (新)")

        # ── 球队 ─────────────────────────────────────────────
        team_sql = """
            INSERT OR IGNORE INTO teams (id, name, name_cn, league_id)
            VALUES (?, ?, ?, ?)
        """
        team_count = 0
        for tid, info in self.teams_cache.items():
            c = self.conn.execute(team_sql, (
                info["id"], info["name"], info.get("name_cn", ""),
                info.get("league_id"),
            ))
            if c.rowcount > 0:
                team_count += 1
        self.stats["teams_count"] = team_count
        print(f"  球队: {team_count} 条 (新)")

        # ── 赛季 ─────────────────────────────────────────────
        season_sql = "INSERT OR IGNORE INTO seasons (id, name, start_year, end_year) VALUES (?, ?, ?, ?)"
        for s in self.stats["seasons"]:
            s_name = str(s).strip()
            if s_name:
                try:
                    sy = int(s_name[:4])
                    self.conn.execute(season_sql, (s_name, s_name, sy, sy + 1))
                except ValueError:
                    pass

        # ── 比赛（去重） ─────────────────────────────────────
        match_sql = """
            INSERT OR IGNORE INTO matches (
                match_id, source, league_id, season, home_team_id, home_team,
                away_team_id, away_team, kickoff_time, home_score, away_score,
                score_display, half_time_score, total_goals, half_full_result,
                ft_result, round, status, home_ranking, away_ranking, collected_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        match_inserted = 0
        match_dupes = 0

        for m in matches:
            c = self.conn.execute(match_sql, (
                m["match_id"], m["source"], m["league_id"], m["season"],
                m["home_team_id"], m["home_team"], m["away_team_id"], m["away_team"],
                m["kickoff_time"], m["home_score"], m["away_score"],
                m["score_display"], m["half_time_score"], m["total_goals"],
                m["half_full_result"], m["ft_result"], m["round"], m["status"],
                m["home_ranking"], m["away_ranking"], m["collected_at"],
            ))
            if c.rowcount > 0:
                match_inserted += 1
            else:
                match_dupes += 1

        self.stats["duplicates"] = match_dupes
        print(f"  比赛: {match_inserted} 条 (新增), {match_dupes} 条 (重复跳过)")

        # ── 欧洲赔率 ─────────────────────────────────────────
        euro_sql = """
            INSERT OR IGNORE INTO odds_europe (
                match_id, bookmaker, odds_type, odds_home, odds_draw, odds_away
            ) VALUES (?, ?, ?, ?, ?, ?)
        """
        euro_count = 0
        for eo in euro_odds:
            c = self.conn.execute(euro_sql, (
                eo.get("match_id", ""), eo["bookmaker"], eo["odds_type"],
                eo.get("odds_home"), eo.get("odds_draw"), eo.get("odds_away"),
            ))
            if c.rowcount > 0:
                euro_count += 1
        self.stats["euro_odds_count"] = euro_count
        print(f"  欧洲赔率: {euro_count} 条")

        # ── 亚盘 ─────────────────────────────────────────────
        asian_sql = """
            INSERT OR IGNORE INTO odds_asian (
                match_id, bookmaker, odds_type, high_water, handicap, low_water
            ) VALUES (?, ?, ?, ?, ?, ?)
        """
        asian_count = 0
        for ao in asian_odds:
            c = self.conn.execute(asian_sql, (
                ao.get("match_id", ""), ao["bookmaker"], ao["odds_type"],
                ao.get("high_water"), ao.get("handicap", ""), ao.get("low_water"),
            ))
            if c.rowcount > 0:
                asian_count += 1
        self.stats["asian_odds_count"] = asian_count
        print(f"  亚盘: {asian_count} 条")

        # ── 大小球 ───────────────────────────────────────────
        ou_sql = """
            INSERT OR IGNORE INTO odds_over_under (
                match_id, bookmaker, odds_type, over_water, handicap, under_water
            ) VALUES (?, ?, ?, ?, ?, ?)
        """
        ou_count = 0
        for ou in over_under_odds:
            c = self.conn.execute(ou_sql, (
                ou.get("match_id", ""), ou["bookmaker"], ou["odds_type"],
                ou.get("over_water"), ou.get("handicap", ""), ou.get("under_water"),
            ))
            if c.rowcount > 0:
                ou_count += 1
        self.stats["over_under_odds_count"] = ou_count
        print(f"  大小球: {ou_count} 条")

        self.conn.commit()

    # ── AI/回测特征计算 ─────────────────────────────────────

    def compute_features(self):
        """计算 AI/回测特征字段"""
        import statistics as stats

        print(f"\n{'='*60}")
        print("计算 AI/回测特征...")
        print(f"{'='*60}")

        # 获取所有比赛及其欧赔
        matches = self.conn.execute("""
            SELECT match_id, home_score, away_score, total_goals,
                   half_time_score, ft_result
            FROM matches
        """).fetchall()

        bet365_odds = {}
        all_euro_odds = {}
        for row in self.conn.execute("""
            SELECT match_id, bookmaker, odds_type, odds_home, odds_draw, odds_away
            FROM odds_europe
        """).fetchall():
            mid, bk, otype, oh, od, oa = row
            if mid not in all_euro_odds:
                all_euro_odds[mid] = {"opening": [], "closing": []}
            key = "opening" if otype == "opening" else "closing"
            all_euro_odds[mid][key].append((bk, oh, od, oa))
            if bk == "Bet365":
                if mid not in bet365_odds:
                    bet365_odds[mid] = {}
                bet365_odds[mid][otype] = (oh, od, oa)

        # 获取亚盘数据
        asian_odds_data = {}
        for row in self.conn.execute("""
            SELECT match_id, bookmaker, odds_type, handicap
            FROM odds_asian WHERE bookmaker = 'Bet365'
        """).fetchall():
            mid, bk, otype, handicap = row
            if mid not in asian_odds_data:
                asian_odds_data[mid] = {}
            asian_odds_data[mid][otype] = handicap

        feature_count = 0
        feature_sql = """
            INSERT OR REPLACE INTO match_features (
                match_id, implied_home_prob, implied_draw_prob, implied_away_prob,
                odds_movement_home, odds_movement_draw, odds_movement_away,
                asian_handicap_open, asian_handicap_close, asian_movement_flag,
                over_under_line, over_prob_implied,
                odds_std_home, odds_std_draw, odds_std_away,
                home_win, draw, away_win, total_goals, over_2_5, both_teams_score
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """

        for match in matches:
            mid, h_score, a_score, t_goals, ht_score, ft_result = match
            odds = bet365_odds.get(mid, {})
            all_odds = all_euro_odds.get(mid, {"opening": [], "closing": []})
            asian = asian_odds_data.get(mid, {})

            # ── 隐含概率 (Bet365 closing) ──────────────────────
            closing = odds.get("closing")
            if closing and all(v is not None for v in closing):
                oh, od, oa = closing
                total_payout = 1/oh + 1/od + 1/oa if oh and od and oa else 1
                implied_home = (1/oh) / total_payout if oh else None
                implied_draw = (1/od) / total_payout if od else None
                implied_away = (1/oa) / total_payout if oa else None
            else:
                implied_home = implied_draw = implied_away = None

            # ── 赔率变化 (Bet365 opening → closing) ──────────
            opening = odds.get("opening")
            if opening and closing and all(v is not None for v in opening + closing):
                o_h, o_d, o_a = opening
                c_h, c_d, c_a = closing
                movement_h = c_h - o_h if o_h and c_h else None
                movement_d = c_d - o_d if o_d and c_d else None
                movement_a = c_a - o_a if o_a and c_a else None
            else:
                movement_h = movement_d = movement_a = None

            # ── 亚盘变化 ─────────────────────────────────────
            asian_open = asian.get("opening", "")
            asian_close = asian.get("closing", "")
            asian_flag = 1 if asian_open != asian_close else 0

            # ── 大小球盘口 ──────────────────────────────────
            ou_line = None
            ou_row = self.conn.execute(
                "SELECT handicap FROM odds_over_under WHERE match_id=? AND bookmaker='Bet365' AND odds_type='closing' LIMIT 1",
                (mid,)
            ).fetchone()
            if ou_row:
                try:
                    ou_line = float(ou_row[0].replace(" ", ""))
                except (ValueError, AttributeError):
                    ou_line = None

            over_prob = None
            ou_odds_row = self.conn.execute(
                "SELECT over_water, under_water FROM odds_over_under WHERE match_id=? AND bookmaker='Bet365' AND odds_type='closing' LIMIT 1",
                (mid,)
            ).fetchone()
            if ou_odds_row and ou_odds_row[0] and ou_odds_row[1]:
                total = 1/ou_odds_row[0] + 1/ou_odds_row[1]
                over_prob = (1/ou_odds_row[0]) / total if total else None

            # ── 赔率离散度 (多家博彩公司 closing 的标准差) ──
            closing_odds = all_odds["closing"]
            if len(closing_odds) >= 3:
                homes = [x[1] for x in closing_odds if x[1] is not None]
                draws = [x[2] for x in closing_odds if x[2] is not None]
                aways = [x[3] for x in closing_odds if x[3] is not None]
                std_h = stats.stdev(homes) if len(homes) >= 2 else None
                std_d = stats.stdev(draws) if len(draws) >= 2 else None
                std_a = stats.stdev(aways) if len(aways) >= 2 else None
            else:
                std_h = std_d = std_a = None

            # ── 标签 ─────────────────────────────────────────
            home_win = 1 if ft_result == "胜" else 0
            draw = 1 if ft_result == "平" else 0
            away_win = 1 if ft_result == "负" else 0
            goals = t_goals if t_goals is not None else (
                (h_score + a_score) if h_score is not None and a_score is not None else None
            )
            over_2_5 = 1 if goals is not None and goals > 2.5 else (0 if goals is not None else 0)
            both_score = 1 if (h_score is not None and a_score is not None
                               and h_score > 0 and a_score > 0) else 0

            self.conn.execute(feature_sql, (
                mid, implied_home, implied_draw, implied_away,
                movement_h, movement_d, movement_a,
                asian_open, asian_close, asian_flag,
                ou_line, over_prob,
                std_h, std_d, std_a,
                home_win, draw, away_win, goals, over_2_5, both_score,
            ))
            feature_count += 1

        self.conn.commit()
        print(f"  特征计算完成: {feature_count} 条")
        return feature_count

    # ── CSV/JSON 导出 ────────────────────────────────────────

    def export_csv(self):
        """导出数据库所有表为 CSV"""
        import csv

        csv_dir = self.processed_dir / "csv"
        csv_dir.mkdir(parents=True, exist_ok=True)

        tables = ["leagues", "teams", "seasons", "matches",
                   "odds_europe", "odds_asian", "odds_over_under",
                   "match_features"]

        for table in tables:
            try:
                cursor = self.conn.execute(f"SELECT * FROM {table}")
                rows = cursor.fetchall()
                if not rows:
                    continue
                cols = [desc[0] for desc in cursor.description]
                filepath = csv_dir / f"{table}.csv"
                with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    writer.writerow(cols)
                    writer.writerows(rows)
                print(f"  CSV: {filepath} ({len(rows)} rows)")
            except Exception as e:
                print(f"  CSV 导出 {table} 失败: {e}")

    def export_json(self):
        """导出数据库所有表为 JSON"""
        json_dir = self.processed_dir / "json"
        json_dir.mkdir(parents=True, exist_ok=True)

        # 导出完整数据包
        data_package = {
            "metadata": {
                "exported_at": datetime.now().isoformat(),
                "source": str(self.excel_path.name),
                "stats": {k: str(v) if isinstance(v, set) else v
                          for k, v in self.stats.items()},
            },
        }

        tables = ["leagues", "teams", "seasons", "matches",
                   "odds_europe", "odds_asian", "odds_over_under",
                   "match_features"]

        for table in tables:
            try:
                cursor = self.conn.execute(f"SELECT * FROM {table}")
                rows = cursor.fetchall()
                cols = [desc[0] for desc in cursor.description]
                data = [dict(zip(cols, row)) for row in rows]

                # 写入单独文件
                filepath = json_dir / f"{table}.json"
                with open(filepath, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2, default=str)
                print(f"  JSON: {filepath} ({len(data)} rows)")

                # 加入数据包（只保留摘要）
                data_package[f"{table}_count"] = len(data)
            except Exception as e:
                print(f"  JSON 导出 {table} 失败: {e}")

        # 写入数据包
        pkg_path = json_dir / "data_package.json"
        with open(pkg_path, "w", encoding="utf-8") as f:
            json.dump(data_package, f, ensure_ascii=False, indent=2, default=str)
        print(f"  JSON: {pkg_path}")

    # ── 质量报告 ─────────────────────────────────────────────

    def generate_report(self):
        """生成数据质量报告"""
        report_lines = []
        report_lines.append("=" * 70)
        report_lines.append("足球历史数据导入 — 数据质量报告")
        report_lines.append(f"生成时间: {datetime.now().isoformat()}")
        report_lines.append(f"源文件: {self.excel_path.name}")
        report_lines.append("=" * 70)

        # 总体统计
        report_lines.append("\n## 总体统计")
        report_lines.append(f"  总行数: {self.stats['total_rows']}")
        report_lines.append(f"  有效比赛: {self.stats['valid_matches']}")
        report_lines.append(f"  跳过空行: {self.stats['skipped_empty']}")
        report_lines.append(f"  重复跳过: {self.stats['duplicates']}")
        report_lines.append(f"  错误数: {len(self.stats['errors'])}")
        report_lines.append(f"  联赛数: {self.stats['leagues_count']}")
        report_lines.append(f"  球队数: {self.stats['teams_count']}")
        report_lines.append(f"  赛季: {sorted(self.stats['seasons'])}")
        report_lines.append(f"  欧洲赔率: {self.stats['euro_odds_count']} 条")
        report_lines.append(f"  亚盘: {self.stats['asian_odds_count']} 条")
        report_lines.append(f"  大小球: {self.stats['over_under_odds_count']} 条")

        # 各 sheet 统计
        report_lines.append("\n## 各 Sheet 统计")
        report_lines.append(f"  {'Sheet':<12} {'行数':<8} {'有效':<8} {'跳过':<8} {'耗时':<10}")
        report_lines.append("  " + "-" * 50)
        for sheet, info in self.stats["per_sheet"].items():
            report_lines.append(
                f"  {sheet:<12} {info['rows']:<8} {info['valid']:<8} "
                f"{info['skipped']:<8} {info['seconds']:.1f}s"
            )

        # 数据库统计
        report_lines.append("\n## 数据库内容统计")
        tables = ["leagues", "teams", "seasons", "matches",
                   "odds_europe", "odds_asian", "odds_over_under",
                   "match_features", "import_logs"]
        for table in tables:
            try:
                cursor = self.conn.execute(f"SELECT COUNT(*) FROM {table}")
                count = cursor.fetchone()[0]
                report_lines.append(f"  {table}: {count} 条")
            except Exception:
                pass

        # 联赛分布
        report_lines.append("\n## 联赛分布")
        try:
            cursor = self.conn.execute("""
                SELECT l.name_cn, COUNT(*) as cnt
                FROM matches m JOIN leagues l ON m.league_id = l.id
                GROUP BY m.league_id ORDER BY cnt DESC
            """)
            for row in cursor.fetchall():
                report_lines.append(f"  {row[0]}: {row[1]} 场比赛")
        except Exception as e:
            report_lines.append(f"  查询失败: {e}")

        # 赛季分布
        report_lines.append("\n## 赛季分布")
        try:
            cursor = self.conn.execute("""
                SELECT season, COUNT(*) as cnt
                FROM matches GROUP BY season ORDER BY season
            """)
            for row in cursor.fetchall():
                report_lines.append(f"  {row[0]}: {row[1]} 场比赛")
        except Exception as e:
            report_lines.append(f"  查询失败: {e}")

        # 赛果分布
        report_lines.append("\n## 赛果分布 (胜平负)")
        try:
            cursor = self.conn.execute("""
                SELECT ft_result, COUNT(*) as cnt
                FROM matches WHERE ft_result != ''
                GROUP BY ft_result ORDER BY cnt DESC
            """)
            for row in cursor.fetchall():
                report_lines.append(f"  {row[0]}: {row[1]} 场")
        except Exception as e:
            report_lines.append(f"  查询失败: {e}")

        # 比分完整性
        report_lines.append("\n## 数据完整性")
        try:
            cursor = self.conn.execute("""
                SELECT
                    COUNT(*) as total,
                    SUM(CASE WHEN home_score IS NOT NULL THEN 1 ELSE 0 END) as has_score,
                    SUM(CASE WHEN ft_result != '' THEN 1 ELSE 0 END) as has_result,
                    SUM(CASE WHEN kickoff_time IS NOT NULL THEN 1 ELSE 0 END) as has_kickoff
                FROM matches
            """)
            row = cursor.fetchone()
            total = row[0] or 0
            report_lines.append(f"  总比赛: {total}")
            report_lines.append(f"  有比分: {row[1]} ({row[1]/total*100:.1f}%)" if total else "  有比分: 0")
            report_lines.append(f"  有赛果: {row[2]} ({row[2]/total*100:.1f}%)" if total else "  有赛果: 0")
            report_lines.append(f"  有开球时间: {row[3]} ({row[3]/total*100:.1f}%)" if total else "  有开球时间: 0")
        except Exception as e:
            report_lines.append(f"  查询失败: {e}")

        # 博彩公司分布
        report_lines.append("\n## 博彩公司覆盖 (欧洲赔率)")
        try:
            cursor = self.conn.execute("""
                SELECT bookmaker, odds_type, COUNT(*) as cnt
                FROM odds_europe GROUP BY bookmaker, odds_type
                ORDER BY bookmaker, odds_type
            """)
            for row in cursor.fetchall():
                report_lines.append(f"  {row[0]} ({row[1]}): {row[2]} 条")
        except Exception as e:
            report_lines.append(f"  查询失败: {e}")

        # 错误详情
        if self.stats["errors"]:
            report_lines.append(f"\n## 错误详情 (前20条)")
            for err in self.stats["errors"][:20]:
                report_lines.append(f"  - {err}")

        report_lines.append(f"\n{'=' * 70}")
        report_lines.append("报告结束")

        report_text = "\n".join(report_lines)

        # 保存报告
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = self.reports_dir / f"import_report_{timestamp}.txt"
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(report_text)

        # 同时保存 JSON 报告
        json_report_path = self.reports_dir / f"import_report_{timestamp}.json"
        json_stats = {k: (sorted(v) if isinstance(v, set) else v)
                      for k, v in self.stats.items()}
        with open(json_report_path, "w", encoding="utf-8") as f:
            json.dump(json_stats, f, ensure_ascii=False, indent=2, default=str)

        print(report_text)
        print(f"\n报告已保存: {report_path}")
        print(f"JSON 报告: {json_report_path}")
        return report_text

    # ── 导入日志 ─────────────────────────────────────────────

    def save_import_log(self, sheet_name: str, total: int, valid: int,
                        skipped: int, dupes: int, duration: float,
                        errors: str = ""):
        """记录导入日志"""
        self.conn.execute(
            """INSERT INTO import_logs
               (source_file, source_sheet, total_rows, valid_rows,
                skipped_rows, duplicate_rows, errors, duration_seconds)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (str(self.excel_path.name), sheet_name, total, valid,
             skipped, dupes, errors[:1000] if errors else "", duration)
        )
        self.conn.commit()

    # ── 主流程 ──────────────────────────────────────────────

    def run(self, skip_csv: bool = False, skip_json: bool = False):
        """执行完整的导入流程"""
        total_start = time.time()

        # 1. 解析 Excel
        matches, euro_odds, asian_odds, over_under_odds = self.parse_excel()

        if not matches:
            print("\n错误: 没有解析到任何比赛数据！")
            return

        # 2. 建立数据库
        print(f"\n建立数据库: {self.db_path}")
        self.setup_database()

        # 3. 关联 match_id 到赔率
        print("关联赔率到比赛...")
        match_id_map = {}
        for m in matches:
            match_id_map[m["match_id"]] = m

        for odds_list in [euro_odds, asian_odds, over_under_odds]:
            for odds in odds_list:
                # 赔率数据从对应 match 获取 match_id
                # 当前赔率数据结构中不包含 match_id，需要关联
                # 我们在 _parse_row 时就已按顺序关联
                pass

        # 4. 写入数据库
        self.save_to_database(matches, euro_odds, asian_odds, over_under_odds)

        # 5. 计算 AI/回测特征
        self.compute_features()

        # 6. 导出 CSV
        if not skip_csv:
            print(f"\n导出 CSV 到 {self.processed_dir / 'csv'}")
            self.export_csv()

        # 7. 导出 JSON
        if not skip_json:
            print(f"\n导出 JSON 到 {self.processed_dir / 'json'}")
            self.export_json()

        # 7. 生成质量报告
        total_elapsed = time.time() - total_start
        print(f"\n{'='*60}")
        print(f"导入完成! 总耗时: {total_elapsed:.1f}s")
        self.generate_report()

        # 8. 数据库优化
        print("\n优化数据库...")
        self.conn.execute("ANALYZE")
        self.conn.execute("PRAGMA optimize")
        self.conn.close()
        print("数据库已关闭。")

        return True


def main():
    import argparse

    parser = argparse.ArgumentParser(description="历史足球数据导入中心")
    parser.add_argument("--excel", type=str, default=None, help="Excel 文件路径")
    parser.add_argument("--skip-csv", action="store_true", help="跳过 CSV 导出")
    parser.add_argument("--skip-json", action="store_true", help="跳过 JSON 导出")
    parser.add_argument("--db", type=str, default=None, help="SQLite 数据库路径")
    args = parser.parse_args()

    importer = HistoricalDataImporter(excel_path=args.excel)

    if args.db:
        importer.db_path = Path(args.db)

    success = importer.run(skip_csv=args.skip_csv, skip_json=args.skip_json)

    if success:
        print("\nData import pipeline completed successfully!")
    else:
        print("\nData import pipeline failed!")
        sys.exit(1)


if __name__ == "__main__":
    main()
